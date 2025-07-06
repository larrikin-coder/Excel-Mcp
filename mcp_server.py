import os
import json
from dotenv import load_dotenv

from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse
import google.generativeai as genai
from openai import AsyncOpenAI
from openpyxl import Workbook, load_workbook

import uvicorn


from write_cell import *
from create_sheet import *

load_dotenv()

api_key = os.getenv("OPENAI_API_KEY")
openai_client = AsyncOpenAI(api_key=api_key)

genai.configure(api_key=os.getenv("GEMINI_API_KEY"))
model = genai.GenerativeModel('gemini-2.5-flash')

class MCPServer:
    def __init__(self, name, description, parameters, function):
        self.name = name
        self.description = description
        self.parameters = parameters
        self.function = function

    def run(self, **kwargs):
        return self.function(**kwargs)

    @staticmethod
    def from_function(fn):
        import inspect
        sig = inspect.signature(fn)
        params = {
            name: {"type": "string", "description": ""}
            for name in sig.parameters
        }
        return MCPServer(
            name=fn.__name__,
            description=fn.__doc__ or "",
            parameters=params,
            function=fn
        )

    def openai_tool(self):
        return {
            "type": "function",
            "function": {
                "name": self.name,
                "description": self.description,
                "parameters": {
                    "type": "object",
                    "properties": self.parameters,
                    "required": list(self.parameters.keys())
                }
            }
        }


class ToolFunction:
    @staticmethod
    def from_function(fn):
        return MCPServer.from_function(fn)


class Middleware:
    def __init__(self, tool_functions):
        self.tool_functions = tool_functions

    async def call(self,client, res):
        try:
            tool_call = res['tool_calls'][0]
            tool_name = tool_call['function']['name']
            tool_args_raw = tool_call['function']['arguments']

            # Handle both string or dict input
            if isinstance(tool_args_raw, str):
                tool_args = json.loads(tool_args_raw)
            elif isinstance(tool_args_raw, dict):
                tool_args = tool_args_raw
            else:
                raise ValueError("Invalid arguments format")

            for tool in self.tool_functions:
                if tool.name == tool_name:
                    return tool.run(**tool_args)

            raise Exception(f"No tool found for {tool_name}")

        except Exception as e:
            print(f"Middleware Error: {e}")
            raise e

    def openai_tools(self):
        return [tool.openai_tool() for tool in self.tool_functions]


# def create_sheet(filepath: str, sheet_name: str):
#     """Create a new sheet, create file if it doesn't exist."""
#     if not os.path.exists(filepath):
#         wb = Workbook()
#         wb.create_sheet(title=sheet_name)
#         default_sheet = wb["Sheet"]
#         wb.remove(default_sheet)
#     else:
#         wb = load_workbook(filepath)
#         wb.create_sheet(title=sheet_name)

#     wb.save(filepath)
#     return {"message": f"Sheet '{sheet_name}' created in '{filepath}'."}

# def write_cell(filepath: str, sheet_name: str, cell: str, value: str):
#     """Write a value to a specific cell. Create sheet if missing. Handle spaces, cases safely."""
#     wb = load_workbook(filepath)

#     clean_sheet_names = {s.strip().lower(): s for s in wb.sheetnames}
#     requested_clean = sheet_name.strip().lower()

#     if requested_clean in clean_sheet_names:
#         ws = wb[clean_sheet_names[requested_clean]]
#     else:
#         ws = wb.create_sheet(title=sheet_name)

#     ws[cell] = value
#     wb.save(filepath)
#     return {"message": f"Value '{value}' written to {sheet_name}:{cell}"}

app = FastAPI()

mcp_handler = Middleware(
    tool_functions=[
        ToolFunction.from_function(write_cell),
        ToolFunction.from_function(create_sheet),
    ]
)



@app.post("/mcp")
async def handle_mcp(request: Request):
    try:
        body = await request.json()
        result = await mcp_handler.call(openai_client,body)
        return JSONResponse(content=result)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})




# @app.post("/chat")
# async def chatgpt(request: Request):
#     data =  await request.json()
#     prompt = data.get('prompt')
    
    
#     default_filepath = "uploaded_file.xlsx"
    
#     response = await openai_client.chat.completions.create(
#         model="gpt-4o",
#         messages=[{
#             "role":"system",
#             "content":(
#                 "You are an Excel control agent."
#                 "You have to use only MCP tools for the prompts give by users."
#                 "Response should only be generated in structred MCPtool calls."
#                 "You have full access of the MCP Tools and can plan accordingly how to call them for solving the user prompts"
#                 f"If a tool requires a 'filepath' argument then use the {default_filepath} "
#                 "Respond with short lines"
#                 "Only tell whether the prompt instructions are successfully executed or not"
#                 "If an error occurs explain the possible causes in a crisp and concise manner"
                
#             )  
#         },{
#             "role":"user",
#             "content": prompt
#         }],
#         tools=mcp_handler.openai_tools(),
#         tool_choice = "auto"
#         )     
#     tool_calls = response.choices[0].message.tool_calls
#     if not tool_calls:
#         return{"error":"No MCP Tools Found in the model."}
    
#     all_responses = []
#     for tool_call_obj in tool_calls:
#         tool_call = tool_call_obj.to_dict()
#         arguments = tool_call["function"]["arguments"]
#         if isinstance(arguments, str):
#             arguments = json.loads(arguments)

#         if "filepath" not in arguments:
#             arguments["filepath"] = default_filepath

#         tool_call["function"]["arguments"] = arguments

#         # Execute tool call
#         mcp_response = await mcp_handler.call(openai_client, {"tool_calls": [tool_call]})
#         all_responses.append(mcp_response)

#     return {"results": all_responses}

import traceback


def extract_json_from_markdown(raw_text: str) -> str:
    """
    Extract JSON from a Markdown-style response like:
    ```json
    { "key": "value" }
    ```
    """
    raw_text = raw_text.strip()
    
    # If the string starts with a markdown block (```json)
    if raw_text.startswith("```"):
        lines = raw_text.splitlines()
        # Remove first and last line (```json and ```)
        if len(lines) >= 3:
            return "\n".join(lines[1:-1])
    
    return raw_text  # Assume it's clean JSON otherwise




@app.post("/chat")
async def chatgpt(request: Request):
    try:
        data = await request.json()
        prompt = data.get("prompt")
        default_filepath = "uploaded_file.xlsx"

        GEMINI_SYSTEM_PROMPT = """
        You are an Excel control agent. You are only allowed to respond with a JSON structure that represents a tool/function call.

        Format:
        {
        "function": {
            "name": "tool_name",
            "arguments": {
            "param1": "value1",
            "param2": "value2",
            "filepath": "uploaded_file.xlsx"
            }
        }
        }

        Only use the available tools: `create_sheet`, `write_cell`.
        You can use the functions concurrently if the prompt requires you to.
        Details:
        - `create_sheet`: needs `filepath` and `sheet_name`
        - `write_cell`: needs `filepath`, `sheet_name`, `cell`, and `value`

        ✅ Do not explain.
        ✅ Do not add extra text.
        ✅ Always return only the JSON structure.

        If any argument is missing from user input, guess reasonable values.
        """
 
        # Format the prompt with system instruction
        formatted_prompt = f"{GEMINI_SYSTEM_PROMPT.strip()}\n\nUser input: {prompt}"

        # Call Gemini
        response = model.generate_content(
            contents=[
                {
                    "role": "user",
                    "parts": [formatted_prompt]
                }
            ]
        )

        reply = response.text.strip()
        reply = extract_json_from_markdown(reply)

        # Try to parse as JSON function call
        try:
            tool_data = json.loads(reply)
            if not isinstance(tool_data, list):
                tool_data = [tool_data]
        except Exception as e:
            return JSONResponse(status_code=500, content={
                "error": f"Failed to parse Gemini response as JSON: {e}",
                "raw_response": reply
            })

        tool_calls = []
        for tool in tool_data:
            if "function" in tool:
                args = tool["function"]["arguments"]
                if "filepath" not in args:
                    args["filepath"] = default_filepath
                tool_calls.append(tool)

        all_responses = []
        for tool_call in tool_calls:
            mcp_response = await mcp_handler.call(None, {"tool_calls": [tool_call]})
            all_responses.append(mcp_response)

        return {"results": all_responses}

    except Exception as e:
        traceback.print_exc()  # <-- This will print the real error in the terminal
        return JSONResponse(status_code=500, content={"error": str(e)})








if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)
