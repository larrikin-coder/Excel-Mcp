from modules import load_workbook, file_to_base64

def write_cell(filepath: str, sheet_name: str, cell: str, value: str):
    """Write a value to a specific cell. Create sheet if missing. Return updated file in base64."""

    try:
        wb = load_workbook(filepath)
    except FileNotFoundError:
        # If file doesn’t exist, create a new one
        from openpyxl import Workbook
        wb = Workbook()
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)
        wb.create_sheet(title=sheet_name)

    clean_sheet_names = {s.strip().lower(): s for s in wb.sheetnames}
    requested_clean = sheet_name.strip().lower()

    if requested_clean in clean_sheet_names:
        ws = wb[clean_sheet_names[requested_clean]]
    else:
        ws = wb.create_sheet(title=sheet_name)

    ws[cell] = value
    wb.save(filepath)

    return {
        "message": f"Value '{value}' written to {sheet_name}:{cell}",
        "file": file_to_base64(filepath)
    }
